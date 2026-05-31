"""Чтение .xlsx/.xlsm через openpyxl (read-only, потоково).

Заголовки берём из первой строки активного листа. Значения отдаём как
есть (числа/даты остаются нативными типами) — конвертация в процессоре.
"""

import io

from openpyxl import load_workbook


def _cell(value):
    """None → '' (как пустая ячейка CSV); строки strip'аем.

    raw-строка сохраняется в JSONField, поэтому несериализуемые типы
    (datetime/date/time и пр.) приводим к строке. int/float/bool — оставляем.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool | int | float):
        return value
    return str(value)


def _workbook(django_file):
    django_file.open("rb")
    try:
        django_file.seek(0)
    except (OSError, ValueError):
        pass
    data = django_file.read()
    # BytesIO — seekable, нужно openpyxl для распаковки zip даже в read_only
    return load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def read_headers(django_file) -> list[str]:
    """Первая строка активного листа как список заголовков."""
    wb = _workbook(django_file)
    try:
        ws = wb.active
        if ws is None:
            return []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [str(_cell(c)).strip() for c in row]
        return []
    finally:
        wb.close()


def read_rows(django_file):
    """Стримить строки как dict {header: value} (по аналогии с csv.DictReader)."""
    wb = _workbook(django_file)
    try:
        ws = wb.active
        if ws is None:
            return
        headers: list[str] | None = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(_cell(c)).strip() for c in row]
                continue
            # пропускаем полностью пустые строки (xlsx часто тянет хвост)
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            data: dict = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                data[header] = _cell(row[i]) if i < len(row) else ""
            yield data
    finally:
        wb.close()
